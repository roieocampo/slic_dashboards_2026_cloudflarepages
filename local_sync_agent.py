"""
SLIC DashBoards Local PC Auto Sync Agent

Run this on the Windows PC that can access the local OneDrive/SharePoint Excel files.
Vercel cannot read C:\\ paths, so this local agent performs the automatic sync.

Required environment variables:
  SUPABASE_URL
  SUPABASE_SERVICE_KEY

Optional for company network SSL/proxy issues:
  SLIC_SSL_VERIFY=false

Install:
  pip install requests openpyxl

Run:
  python local_sync_agent.py
"""
import os
import sys
import re
import io
import json
import time
import uuid
import hashlib
import string
import urllib.parse
from datetime import datetime, timezone
from decimal import Decimal

import requests
from openpyxl import load_workbook

try:
    sys.stdout.reconfigure(line_buffering=True)
except Exception:
    pass

SUPABASE_URL = os.getenv("SUPABASE_URL", "").rstrip("/")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")
ROWS_TABLE = os.getenv("SLIC_ROWS_TABLE", "slic_dashboard_rows")
META_TABLE = os.getenv("SLIC_META_TABLE", "slic_sheet_meta")
LOG_TABLE = os.getenv("SLIC_LOG_TABLE", "slic_sync_log")
SETTINGS_TABLE = os.getenv("SLIC_SETTINGS_TABLE", "slic_sync_settings")
PAGE_SIZE = 500
STOP_AFTER_BLANK_ROWS = int(os.getenv("SLIC_STOP_AFTER_BLANK_ROWS", "500"))
SSL_VERIFY = os.getenv("SLIC_SSL_VERIFY", "true").strip().lower() not in {"0", "false", "no", "off"}
if not SSL_VERIFY:
    try:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    except Exception:
        pass

SHEETS = {
    "LTX": {
        "sheet_name": "3Q26_LTX Updated_MONITORING",
        "header_row": 2,
        "data_start_row": 3,
        "columns": [1, 2, 5, 7, 13, 14, 15, 16, 18, 19],
        "default_path": r"C:\Users\locampo3\OneDrive - Analog Devices, Inc\Ramilo, Kim Jonas's files - SLIC_Sharepoint\LTX Sample weekly.xlsx",
    },
    "ETS": {
        "sheet_name": "3Q26_ETS Updated_MONITORING",
        "header_row": 2,
        "data_start_row": 3,
        "columns": [1, 2, 5, 7, 13, 14, 15, 16, 18, 19],
        "default_path": r"C:\Users\locampo3\OneDrive - Analog Devices, Inc\Ramilo, Kim Jonas's files - SLIC_Sharepoint\SLIC_Activity_Monitoring.xlsm",
    },
}
SELECTED_SOURCE_COLUMNS = [1, 2, 5, 7, 13, 14, 15, 16, 18, 19]
MAX_COLS = len(SELECTED_SOURCE_COLUMNS)


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def q(value):
    return urllib.parse.quote(str(value), safe="")


def headers(prefer=None):
    if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
        raise RuntimeError("Set SUPABASE_URL and SUPABASE_SERVICE_KEY environment variables first.")
    h = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
    }
    if prefer:
        h["Prefer"] = prefer
    return h


def sb_request(method, endpoint, body=None, prefer=None, timeout=90):
    url = f"{SUPABASE_URL}/rest/v1/{endpoint.lstrip('/')}"
    r = requests.request(method, url, headers=headers(prefer), json=body, timeout=timeout, verify=SSL_VERIFY)
    if r.status_code >= 400:
        raise RuntimeError(f"Supabase {method} failed {r.status_code}: {r.text[:500]}")
    if r.status_code == 204 or not r.text:
        return None
    return r.json()


def sb_upsert(table, rows, conflict_cols, timeout=90):
    if not rows:
        return None
    endpoint = f"{table}?on_conflict={q(conflict_cols)}"
    return sb_request("POST", endpoint, rows, prefer="resolution=merge-duplicates,return=minimal", timeout=timeout)


def jsonable(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, Decimal):
        return float(value)
    return value


def normalize_key(text):
    text = "" if text is None else str(text).strip()
    return re.sub(r"\s+", " ", text)


def unique_headers(raw_headers):
    headers_out, seen = [], {}
    for idx, h in enumerate(raw_headers[:MAX_COLS]):
        letter = string.ascii_uppercase[idx]
        base = normalize_key(h) or f"Column {letter}"
        if base in seen:
            seen[base] += 1
            base = f"{base} ({seen[base]})"
        else:
            seen[base] = 1
        headers_out.append(base)
    while len(headers_out) < MAX_COLS:
        headers_out.append(f"Column {string.ascii_uppercase[len(headers_out)]}")
    return headers_out


def display_header(source_col, raw_header):
    forced = {
        1: "Partname",
        13: "Enrollment Status",
        14: "Cloud Enrollment Remarks",
        15: "TRS Updates",
        16: "PE Updates",
        18: "Final Remarks / Status",
        19: "Week Completion",
    }
    return forced.get(source_col, normalize_key(raw_header) or f"Column {string.ascii_uppercase[source_col - 1]}")


def row_hash(row_values):
    return hashlib.sha256(json.dumps(row_values, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")).hexdigest()


def norm_key(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def header_index(headers_list, names):
    targets = [norm_key(n) for n in names]
    normalized = [norm_key(h) for h in headers_list]
    for target in targets:
        for idx, h in enumerate(normalized):
            if h == target:
                return idx
    for target in targets:
        for idx, h in enumerate(normalized):
            if target and (target in h or h in target):
                return idx
    return None


def status_completed(value):
    v = str(value or "").strip().lower()
    return v in {"completed", "complete", "done", "closed", "finished"} or "complete" in v


def parse_file(sheet_key, path):
    sheet_def = SHEETS[sheet_key]
    sheet_name = sheet_def["sheet_name"]
    if not os.path.exists(path):
        raise RuntimeError(f"{sheet_key} file not found: {path}")

    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {sheet_key}: loading workbook...", flush=True)
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    try:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Sheet not found for {sheet_key}: {sheet_name}")
        ws = wb[sheet_name]
        selected_cols = sheet_def["columns"]
        max_col = max(selected_cols)
        selected_indexes = [col - 1 for col in selected_cols]

        # Read header in one row operation instead of repeated ws.cell calls.
        header_tuple = next(ws.iter_rows(min_row=sheet_def["header_row"], max_row=sheet_def["header_row"], max_col=max_col, values_only=True))
        raw_headers = [jsonable(header_tuple[i] if i < len(header_tuple) else "") for i in selected_indexes]
        headers_list = unique_headers([display_header(selected_cols[i], raw_headers[i]) for i in range(len(selected_cols))])
        status_idx = header_index(headers_list, ["PE Updates", "Status"])

        rows = []
        blank_streak = 0
        scanned = 0
        max_row = ws.max_row or sheet_def["data_start_row"]
        print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] Reading {sheet_key}: sheet '{sheet_name}', Excel max row {max_row:,}...", flush=True)
        print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {sheet_key}: fast row scan started...", flush=True)

        for offset, row_tuple in enumerate(ws.iter_rows(min_row=sheet_def["data_start_row"], max_row=max_row, max_col=max_col, values_only=True), start=0):
            source_row = sheet_def["data_start_row"] + offset
            scanned += 1
            values = [jsonable(row_tuple[i] if i < len(row_tuple) else "") for i in selected_indexes]
            while len(values) < MAX_COLS:
                values.append("")

            if all(str(v or "").strip() == "" for v in values):
                blank_streak += 1
                if blank_streak >= STOP_AFTER_BLANK_ROWS:
                    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {sheet_key}: stopped after {STOP_AFTER_BLANK_ROWS:,} consecutive blank rows at Excel row {source_row:,}.", flush=True)
                    break
                if scanned % 1000 == 0:
                    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {sheet_key}: scanned {scanned:,} rows, parsed {len(rows):,} nonblank rows...", flush=True)
                continue

            blank_streak = 0
            data = {headers_list[i]: values[i] for i in range(MAX_COLS)}
            status_value = values[status_idx] if status_idx is not None else ""
            rows.append({
                "sheet_key": sheet_key,
                "sheet_name": sheet_name,
                "source_row": source_row,
                "row_values": values,
                "data": data,
                "row_hash": row_hash(values),
                "status_value": str(status_value or "").strip(),
                "completed_flag": status_completed(status_value),
            })
            if len(rows) % 2500 == 0:
                print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {sheet_key}: parsed {len(rows):,} nonblank rows...", flush=True)
            elif scanned % 1000 == 0:
                print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {sheet_key}: scanned {scanned:,} rows, parsed {len(rows):,} nonblank rows...", flush=True)

        print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {sheet_key}: finished parsing {len(rows):,} nonblank rows.", flush=True)
        return rows, headers_list
    finally:
        try:
            wb.close()
        except Exception:
            pass


def sync_sheet(sheet_key, path, sync_id):
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {sheet_key}: opening file: {path}", flush=True)
    rows, headers_list = parse_file(sheet_key, path)
    stamp = now_iso()
    for row in rows:
        row["sync_id"] = sync_id
        row["synced_at"] = stamp
    total_batches = (len(rows) + PAGE_SIZE - 1) // PAGE_SIZE if rows else 0
    for batch_no, i in enumerate(range(0, len(rows), PAGE_SIZE), start=1):
        print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {sheet_key}: uploading batch {batch_no:,}/{total_batches:,} ({min(i+PAGE_SIZE, len(rows)):,}/{len(rows):,})...", flush=True)
        sb_upsert(ROWS_TABLE, rows[i:i+PAGE_SIZE], "sheet_key,source_row", timeout=180)
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {sheet_key}: removing old rows...", flush=True)
    sb_request("DELETE", f"{ROWS_TABLE}?sheet_key=eq.{q(sheet_key)}&sync_id=neq.{q(sync_id)}", prefer="return=minimal", timeout=180)
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {sheet_key}: updating sync metadata...", flush=True)
    sb_upsert(META_TABLE, [{
        "sheet_key": sheet_key,
        "sheet_name": SHEETS[sheet_key]["sheet_name"],
        "headers": headers_list,
        "last_sync_id": sync_id,
        "row_count": len(rows),
        "updated_at": now_iso(),
    }], "sheet_key", timeout=180)
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {sheet_key}: done.", flush=True)
    return len(rows)


def get_settings():
    rows = sb_request("GET", f"{SETTINGS_TABLE}?id=eq.1&select=*&limit=1") or []
    settings = {
        "enabled": False,
        "schedule_type": "daily",
        "daily_time": "00:00",
        "interval_minutes": 60,
        "ltx_path": SHEETS["LTX"]["default_path"],
        "ets_path": SHEETS["ETS"]["default_path"],
    }
    if rows:
        settings.update(rows[0])
    return settings


def settings_signature(settings):
    parts = [
        str(bool(settings.get("enabled"))),
        str(settings.get("schedule_type") or "daily"),
        str(settings.get("daily_time") or "00:00"),
        str(settings.get("interval_minutes") or 60),
        str(settings.get("ltx_path") or SHEETS["LTX"]["default_path"]),
        str(settings.get("ets_path") or SHEETS["ETS"]["default_path"]),
    ]
    return "|".join(parts)


def print_settings(settings):
    enabled = bool(settings.get("enabled"))
    mode = settings.get("schedule_type") or "daily"
    if not enabled:
        print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] Auto sync is DISABLED in dashboard settings.")
        return
    if mode == "interval":
        interval = max(5, int(settings.get("interval_minutes") or 60))
        print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] Auto sync ENABLED: every {interval} minute(s).")
    else:
        print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] Auto sync ENABLED: daily at {settings.get('daily_time') or '00:00'}.")
    print(f"  LTX path: {settings.get('ltx_path') or SHEETS['LTX']['default_path']}")
    print(f"  ETS path: {settings.get('ets_path') or SHEETS['ETS']['default_path']}")


def scheduler_decision(settings, state):
    """Return True when it is time to run.

    Interval mode now runs immediately once after the agent starts or after settings change,
    then repeats every N minutes. This avoids the old behavior where it waited for a
    wall-clock minute divisible by N and looked like it was doing nothing.
    """
    if not settings.get("enabled"):
        state["last_interval_run"] = None
        return False

    now = datetime.now()
    mode = settings.get("schedule_type") or "daily"
    if mode == "interval":
        interval_seconds = max(5, int(settings.get("interval_minutes") or 60)) * 60
        last = state.get("last_interval_run")
        if last is None or (time.time() - last) >= interval_seconds:
            state["last_interval_run"] = time.time()
            return True
        return False

    # Daily/specific time mode. Run once per day when the local PC time reaches HH:MM.
    target = settings.get("daily_time") or "00:00"
    today_key = now.strftime("%Y-%m-%d") + " " + target
    if now.strftime("%H:%M") == target and state.get("last_daily_run") != today_key:
        state["last_daily_run"] = today_key
        return True
    return False

def sync_once(settings):
    sync_id = str(uuid.uuid4())
    started = now_iso()
    counts = {"LTX": 0, "ETS": 0}
    counts["LTX"] = sync_sheet("LTX", settings.get("ltx_path") or SHEETS["LTX"]["default_path"], sync_id)
    counts["ETS"] = sync_sheet("ETS", settings.get("ets_path") or SHEETS["ETS"]["default_path"], sync_id)
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] Writing final sync log...", flush=True)
    sb_request("POST", LOG_TABLE, [{
        "sync_id": sync_id,
        "status": "success",
        "message": "Local PC auto sync completed",
        "ets_count": counts["ETS"],
        "ltx_count": counts["LTX"],
        "started_at": started,
        "finished_at": now_iso(),
    }], prefer="return=minimal")
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] Sync completed: LTX={counts['LTX']:,}, ETS={counts['ETS']:,}")


def main():
    print("SLIC DashBoards Local PC Auto Sync Agent started.")
    print(f"SSL certificate verification: {'ON' if SSL_VERIFY else 'OFF'}")
    print("Press Ctrl+C to stop.")
    state = {"signature": None, "last_interval_run": None, "last_daily_run": None}
    while True:
        try:
            settings = get_settings()
            sig = settings_signature(settings)
            if sig != state.get("signature"):
                state["signature"] = sig
                state["last_interval_run"] = None  # run immediately after enabling/changing interval settings
                print_settings(settings)
            if scheduler_decision(settings, state):
                print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] Auto sync started...")
                try:
                    sync_once(settings)
                except Exception as sync_ex:
                    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] SYNC ERROR: {sync_ex}")
        except KeyboardInterrupt:
            print("Stopped.")
            break
        except Exception as ex:
            print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] ERROR: {ex}")
        time.sleep(10)


if __name__ == "__main__":
    main()
